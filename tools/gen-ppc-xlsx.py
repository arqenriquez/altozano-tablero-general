# -*- coding: utf-8 -*-
"""Genera los Excel de metas del módulo PPC (plantilla + semana 05 de ejemplo)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

BASE = os.path.join(os.path.dirname(__file__), "..", "data", "ppc")
HEADERS = ["Actividad", "Responsable", "Lote"]

ACCENT = "2F5D54"
SOFT = "E7EFED"

def estilizar(ws):
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor=ACCENT)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = border
    ws.column_dimensions["A"].width = 58
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 14
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

def escribir(ws, filas):
    thin = Side(style="thin", color="ECECEC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ri, fila in enumerate(filas, start=2):
        for ci, val in enumerate(fila, start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=(ci == 1))
            c.border = border
            if ri % 2 == 0:
                c.fill = PatternFill("solid", fgColor="FAFAF8")

# ---- Plantilla ----
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Metas"
estilizar(ws)
escribir(ws, [
    ["Ejemplo: Colado de losa de cimentación L02", "Ing. Luis Santacruz", "L02"],
    ["Ejemplo: Suministro de acero de la semana", "Metta Admin.", "General"],
    ["", "", ""],
    ["", "", ""],
])
wb.save(os.path.join(BASE, "_plantilla-metas.xlsx"))

# ---- Semana 05 (en curso) ----
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Metas Semana 05"
estilizar(ws)
escribir(ws, [
    ["Colado de losa de cimentación L02", "Ing. Luis Santacruz", "L02"],
    ["Armado de castillos planta baja L03", "Ing. Marco De la Cruz", "L03"],
    ["Excavación de cepas L22", "Ing. Luis Santacruz", "L22"],
    ["Instalación hidrosanitaria en cimentación L04", "Ing. Marco De la Cruz", "L04"],
    ["Relleno y compactación L03", "Ing. Luis Santacruz", "L03"],
    ["Trazo y nivelación L20", "Ing. Marco De la Cruz", "L20"],
    ["Suministro de acero de la semana", "Metta Admin.", "General"],
    ["Limpieza y orden en obra", "Metta Admin.", "General"],
])
wb.save(os.path.join(BASE, "metas", "semana-05.xlsx"))

print("Excel generados:")
print(" -", os.path.normpath(os.path.join(BASE, "_plantilla-metas.xlsx")))
print(" -", os.path.normpath(os.path.join(BASE, "metas", "semana-05.xlsx")))
